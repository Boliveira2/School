from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# Definir artigos e preços
produtos = {
    "Menu 1": "10€",
    "Menu 2": "10€",
    "Menu Criança": "5€",
    "Sardinha": "2€",
    "Broa": "0,50€",
    "Sopa": "2,5€",
    "Bebida": "1,5€",
    "Café": "1€",
    "Bifanas": "3€",
    "Panado": "3€",
    "Cachorro": "3€",
    "Sopa + água (0,50L)": "1€"
}

# Quantidade de senhas a gerar por artigo (podes editar aqui)
quantidades = {
    "Menu 1": 50,
    "Menu 2": 50,
    "Menu Criança": 30,
    "Sardinha": 40,
    "Broa": 40,
    "Sopa": 50,
    "Bebida": 60,
    "Café": 50,
    "Bifanas": 40,
    "Panado": 40,
    "Cachorro": 40,
    "Sopa + água (0,50L)": 30
}

# Setup Excel
wb = Workbook()
ws = wb.active
ws.title = "Senhas Venda Festa"

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
fonte_titulo = Font(bold=True, size=11)
fonte_normal = Font(size=10)

check_box = chr(0x2610)  # caixa vazia

def criar_senha(linha_inicio, col_inicio, produto, preco):
    cell = ws.cell(row=linha_inicio, column=col_inicio)
    texto = (
        f"Festa Final de Ano 2024/2025\n🎫 SENHA DE VENDA\n\n"
        f"Produto: {produto}\n"
        f"Preço: {preco}\n\n"
        f"{check_box}"
    )
    cell.value = texto
    cell.alignment = alinhamento
    cell.font = fonte_normal
    cell.border = border
    ws.merge_cells(
        start_row=linha_inicio, start_column=col_inicio,
        end_row=linha_inicio + 4, end_column=col_inicio + 1
    )

# Configuração de layout
senhas_por_linha = 4
linha_altura = 5  # linhas usadas por cada senha

i = 0
for produto, preco in produtos.items():
    qtd = quantidades.get(produto, 0)
    for _ in range(qtd):
        linha_bloco = (i // senhas_por_linha) * linha_altura + 1
        coluna_bloco = (i % senhas_por_linha) * 2 + 1
        criar_senha(linha_bloco, coluna_bloco, produto, preco)
        i += 1

# Ajustar colunas e linhas
from openpyxl.utils import get_column_letter

for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 20

for row_idx in range(1, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 55

# Guardar ficheiro
wb.save("senhas_venda_festa.xlsx")
print("Senhas individuais para venda geradas no ficheiro: senhas_venda_festa.xlsx")
