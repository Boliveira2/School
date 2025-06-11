import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# Carregar dados do Excel
ficheiro_excel = "reservas.xlsx"
sheet_name = 0

df = pd.read_excel(ficheiro_excel, sheet_name=sheet_name)

# Limpeza dos nomes das colunas
df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip()

# Colunas esperadas
col_nome = "Nome completo do Aluno"
col_turma = "Turma"
col_menu3 = "Quantidade de Menu Criança - 5€"  # Nome exato da coluna para Menu 3 no Excel

produtos_colunas = {
    "Menu 1": "Quantidade de Menu 1 - 10€",
    "Menu 2": "Quantidade de Menu 2 - 10€",
    "Menu Criança": "Quantidade de Menu Criança - 5€",
    "Sardinha": "Sardinha – 2€",
    "Broa": "Broa – 0,50€",
    "Sopa": "Sopa – 2,5€",
    "Bebida": "Bebida – 1,5€",
    "Café": "Café – 1€",
    "Bifanas": "Bifanas – 3€",
    "Panado": "Panado – 3€",
    "Cachorro": "Cachorro – 3€",
    "Sopa + água": "Sopa + água (0,50L) – 1€"
}

# Inicializar workbook
wb = Workbook()
ws = wb.active
ws.title = "Senhas"

# Estilo base
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
fonte_titulo = Font(bold=True, size=11)
fonte_normal = Font(size=9)

# Símbolo quadrado Unicode
check_box = chr(0x2610)

# Função para criar senha formatada
def criar_senha(linha_inicio, col_inicio, nome, turma, produtos):
    cell = ws.cell(row=linha_inicio, column=col_inicio)
    texto = f"Festa Final de Ano 2024/2025\n🎫 SENHA DE RESERVA\n\nNome: {nome}\nTurma: {turma}\n"
    total = 0
    for produto, qtd in produtos.items():
        try:
            qtd_num = int(qtd)
        except (ValueError, TypeError):
            qtd_num = 0
        if qtd_num > 0:
            total += qtd_num
            texto += f"{produto}: {check_box * qtd_num}\n"
    texto += f"\nTotal reservado: {total} produto(s)"
    cell.value = texto
    cell.alignment = alinhamento
    cell.font = fonte_normal
    cell.border = border
    ws.merge_cells(start_row=linha_inicio, start_column=col_inicio,
                   end_row=linha_inicio + 4, end_column=col_inicio + 1)

# Criar senhas
senhas_por_linha = 4
linhas_por_pagina = 20
senhas_por_pagina = senhas_por_linha * (linhas_por_pagina // 5)
i = 0

for _, row in df.iterrows():
    nome = row[col_nome]
    turma = row[col_turma]
    produtos = {}
    for produto, coluna in produtos_colunas.items():
        valor = row.get(coluna, 0)
        if pd.isnull(valor):
            valor = 0
        try:
            valor = int(valor)
        except (ValueError, TypeError):
            valor = 0
        produtos[produto] = valor
    
    # Considerar Menu 3 como 1 unidade se a célula for "Sim"
    valor_menu3 = row.get(col_menu3, "")
    if isinstance(valor_menu3, str) and valor_menu3.strip().lower() == "sim":
        produtos["Menu 3"] = 1
    else:
        produtos["Menu 3"] = 0

    linha_bloco = (i // senhas_por_linha) * 5 + 1
    coluna_bloco = (i % senhas_por_linha) * 2 + 1
    criar_senha(linha_bloco, coluna_bloco, nome, turma, produtos)
    i += 1

# Ajustar tamanhos das colunas e linhas
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 25

for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = 60

# Guardar ficheiro
wb.save("senhas_reservas_formatadas.xlsx")
print("Senhas geradas no ficheiro: senhas_reservas_formatadas.xlsx")
