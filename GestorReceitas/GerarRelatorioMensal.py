import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

# Função para gerar os caminhos corretos dos ficheiros com base no mês
def carregar_ficheiros(mes):
    caminhos = {
        "caf_acolhimento": os.path.join(mes, 'CAF.xlsx'),
        "caf_prolongamento": os.path.join(mes, 'CAF.xlsx'),
        "danca": os.path.join(mes, 'Danca.xlsx'),
        "lanche": os.path.join(mes, 'Lanche.xlsx'),
        "karate": os.path.join(mes, 'Karate.xlsx'),
        "recebimentos": os.path.join(mes, 'recebimentosnumerario.xlsx'),
        "recebimentos_transf": os.path.join(mes, 'transferenciasTratado.xlsx')
    }

    # Verificar se os arquivos existem
    for nome, caminho in caminhos.items():
        if not os.path.exists(caminho):
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    # Carregar os dados
    caf_acolhimento = pd.read_excel(caminhos["caf_acolhimento"], sheet_name='Acolhimento')
    caf_prolongamento = pd.read_excel(caminhos["caf_prolongamento"], sheet_name='Prolongamento')
    danca = pd.read_excel(caminhos["danca"])
    lanche = pd.read_excel(caminhos["lanche"])
    karate = pd.read_excel(caminhos["karate"])
    recebimentos = pd.read_excel(caminhos["recebimentos"])
    recebimentos_transf = pd.read_excel(caminhos["recebimentos_transf"])

    return caf_acolhimento, caf_prolongamento, danca, lanche, recebimentos, recebimentos_transf

# Função para limpar e ajustar os valores das colunas
def ajustar_colunas(df):
    df.columns = df.columns.str.strip()  # Limpar espaços extras nas colunas
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()  # Limpar espaços nas células do tipo string
    return df  

def calcular_custo(nr_dias, preco_unitario):
    return min(nr_dias * 2, preco_unitario)

def obter_valor_recebido_numerario(contribuinte, recebimentos):
    # Remove espaços extras dos nomes das colunas e garante que o contribuinte seja uma string
    recebimentos.columns = recebimentos.columns.str.strip()
    recebimentos['Contribuinte'] = recebimentos['Contribuinte'].astype(str).str.strip()

    # Procura pelo contribuinte no DataFrame
    aluno_recebimentos = recebimentos[recebimentos['Contribuinte'] == str(contribuinte).strip()]
    
    # Se o contribuinte não for encontrado, retorna 0
    if aluno_recebimentos.empty:
        print(f"Nenhum registro encontrado para o contribuinte: {contribuinte}")
        return 0

    # Substitui NaN por 0 nos valores de CAF, Lanche, Dança e Cota, e soma os valores para o contribuinte
    valor_recebido_num = aluno_recebimentos[['CAF', 'Lanche', 'Dança', 'Cota']].fillna(0).sum(axis=1).values[0]
    
    return valor_recebido_num


def obter_valor_recebido_transf(contribuinte, recebimentos):
    """
    Esta função encontra o valor recebido para um determinado contribuinte no DataFrame de recebimentos.
    Se houver múltiplos contribuintes na coluna 'Contribuinte', o valor do crédito é dividido pelo número de contribuintes.
    """
    # Filtrar os recebimentos que correspondem ao contribuinte fornecido
    recebimento_match = recebimentos[recebimentos['Contribuinte'].str.contains(str(contribuinte), na=False)]
    
    # Se houver múltiplos registros para o contribuinte, ajustamos o valor conforme o número de contribuintes
    if not recebimento_match.empty:
        credito_total = recebimento_match['Crédito'].sum()  # Somar o crédito total para o contribuinte
        
        # Verificar se há múltiplos contribuintes
        contribuintes = recebimento_match['Contribuinte'].iloc[0].split(',')  # Dividir o campo "Contribuinte" em uma lista
        num_contribuintes = len(contribuintes)  # Contar quantos contribuintes existem
        
        # Dividir o crédito total pelo número de contribuintes se houver mais de um
        credito_ajustado = credito_total / num_contribuintes if num_contribuintes > 1 else credito_total
        
        return credito_ajustado
    else:
        return 0  # Se não encontrar o contribuinte, retornar 0




# Função para calcular número de dias de acolhimento
def calcular_nr_dias_acolhimento(contribuinte, caf_acolhimento):
    # Remove espaços extras dos nomes das colunas e garante que o contribuinte seja uma string
    caf_acolhimento.columns = caf_acolhimento.columns.str.strip()
    caf_acolhimento['Contribuinte'] = caf_acolhimento['Contribuinte'].astype(str).str.strip()

    # Procura pelo contribuinte no DataFrame
    aluno_acolhimento = caf_acolhimento[caf_acolhimento['Contribuinte'] == str(contribuinte).strip()]
    
    # Se o contribuinte não for encontrado, retorna 0
    if aluno_acolhimento.empty:
        print(f"Nenhum registro encontrado para o contribuinte: {contribuinte}")
        return 0

    #print(f"Registros encontrados para {contribuinte}: {aluno_acolhimento}")
    
    # Substitui "falta" por 0, preenche NaN com 0, e soma os valores nas colunas de dias
    #return pd.to_numeric(aluno_acolhimento.iloc[:, 2:].replace('falta', 0).fillna(0)).sum(axis=1).values[0]

    return aluno_acolhimento.iloc[:, 2:].replace('falta', 0).fillna(0).sum(axis=1).values[0]

# Função para calcular número de dias de prolongamento
def calcular_nr_dias_prolongamento(contribuinte, caf_prolongamento):
    # Remove espaços extras dos nomes das colunas e garante que o contribuinte seja uma string
    caf_prolongamento.columns = caf_prolongamento.columns.str.strip()
    caf_prolongamento['Contribuinte'] = caf_prolongamento['Contribuinte'].astype(str).str.strip()

    # Procura pelo contribuinte no DataFrame
    aluno_prolongamento = caf_prolongamento[caf_prolongamento['Contribuinte'] == str(contribuinte).strip()]
    
    # Se o contribuinte não for encontrado, retorna 0
    if aluno_prolongamento.empty:
        print(f"Nenhum registro encontrado para o contribuinte: {contribuinte}")
        return 0

    #print(f"Registros encontrados para {contribuinte}: {aluno_prolongamento}")
    
    # Substitui "falta" por 0, preenche NaN com 0, e soma os valores nas colunas de dias
    #return pd.to_numeric(aluno_prolongamento.iloc[:, 2:].replace('falta', 0).fillna(0)).sum(axis=1).values[0]

    return aluno_prolongamento.iloc[:, 2:].replace('falta', 0).fillna(0).sum(axis=1).values[0]

# Função para calcular o preço de Dança
def calcular_preco_danca(contribuinte, danca, precos, mes, associado):
    # Remove espaços extras dos nomes das colunas e garante que o contribuinte seja uma string
    danca.columns = danca.columns.str.strip()
    danca['Contribuinte'] = danca['Contribuinte'].astype(str).str.strip()

    # Procura pelo contribuinte no DataFrame
    aluno_danca = danca[danca['Contribuinte'] == str(contribuinte).strip()]
    
    # Se o aluno frequenta dança, retorna o preço
    if not aluno_danca.empty and aluno_danca['Frequenta'].notna().any() and (aluno_danca['Frequenta'] != 0).any():
        if associado == 0:
            return precos[precos['Mês'] == mes]['Preço Dança'].values[0]
        else:
            return precos[precos['Mês'] == mes]['Preço Dança Associado'].values[0]
    
    return 0

# Função para calcular o preço de Karate
def calcular_preco_karate(contribuinte, karate, precos, mes, associado):
    # Remove espaços extras dos nomes das colunas e garante que o contribuinte seja uma string
    karate.columns = karate.columns.str.strip()
    karate['Contribuinte'] = karate['Contribuinte'].astype(str).str.strip()

    # Procura pelo contribuinte no DataFrame
    aluno_karate = karate[karate['Contribuinte'] == str(contribuinte).strip()]
    
    # Se o aluno frequenta dança, retorna o preço
    if not aluno_karate.empty and aluno_karate['Frequenta'].notna().any() and (aluno_karate['Frequenta'] != 0).any():
        if associado == 0:
            return precos[precos['Mês'] == mes]['Preço Karate'].values[0]
        else:
            return precos[precos['Mês'] == mes]['Preço Karate Associado'].values[0]
    
    return 0

# Função para calcular o preço de Lanche
def calcular_preco_lanche(contribuinte, lanche, precos, mes, associado):
    # Remove espaços extras dos nomes das colunas e garante que o contribuinte seja uma string
    lanche.columns = lanche.columns.str.strip()
    lanche['Contribuinte'] = lanche['Contribuinte'].astype(str).str.strip()

    # Procura pelo contribuinte no DataFrame
    aluno_lanche = lanche[lanche['Contribuinte'] == str(contribuinte).strip()]
    
    # Se o aluno frequenta lanche, retorna o preço
    if not aluno_lanche.empty and aluno_lanche['Frequenta'].notna().any() and (aluno_lanche['Frequenta'] != 0).any():
        if associado == 0:
            return precos[precos['Mês'] == mes]['Preço Lanche'].values[0]
        else:
            return precos[precos['Mês'] == mes]['Preço Lanche Associado'].values[0]
    
    return 0


def calcular_preco_caf(contribuinte, mes, caf_acolhimento, caf_prolongamento, precos, associado):
    nr_acolhimento = calcular_nr_dias_acolhimento(contribuinte, caf_acolhimento)
    nr_prolongamento = calcular_nr_dias_prolongamento(contribuinte, caf_prolongamento)
    
    precos['Mês'] = precos['Mês'].str.strip().str.lower()
    mes = mes.strip().lower()
    
   
    if (associado == 0):
        preco_acolhimento = precos[precos['Mês'] == mes]['Preço CAF Acolhimento'].values[0]
        preco_prolongamento = precos[precos['Mês'] == mes]['Preço CAF Prolongamento'].values[0]
        preco_caf = precos[precos['Mês'] == mes]['Preço CAF'].values[0]
        
        custo_acolhimento = calcular_custo(nr_acolhimento, preco_acolhimento)
        custo_prolongamento = calcular_custo(nr_prolongamento, preco_prolongamento)
    else:
        preco_acolhimento = precos[precos['Mês'] == mes]['Preço CAF Acolhimento Associado'].values[0]
        preco_prolongamento = precos[precos['Mês'] == mes]['Preço CAF Prolongamento Associado'].values[0]
        preco_caf = precos[precos['Mês'] == mes]['Preço CAF Associado'].values[0]
        
        custo_acolhimento = calcular_custo(nr_acolhimento, preco_acolhimento)
        custo_prolongamento = calcular_custo(nr_prolongamento, preco_prolongamento)

    return min(custo_acolhimento + custo_prolongamento, preco_caf)


# Geração de relatório mensal
def gerar_relatorioMensal(mes):
    try:
        # Debug: Verifique se a função está recebendo o parâmetro corretamente
        print(f"Gerando relatório para o mês: {mes}")
        # Carregar todos os ficheiros necessários
        caf_acolhimento, caf_prolongamento, danca, lanche, karate, recebimentos, recebimentos_transf = carregar_ficheiros(mes)
    
        alunos = pd.read_csv('InputFiles/alunos.csv', sep=';')
        precos = pd.read_csv('InputFiles/precos.csv', sep=';')
    
        dados_saida = []
        mes_anterior = obter_mes_anterior(mes)
        saldo_anterior = 0
        df_anterior = pd.DataFrame()
    
        if mes_anterior:
            caminho_relatorio_anterior = os.path.join(mes_anterior, f'relatorioMensal_{mes_anterior}.xlsx')
            if os.path.exists(caminho_relatorio_anterior):
                df_anterior = pd.read_excel(caminho_relatorio_anterior)
                saldo_anterior = df_anterior['Saldo'].sum() if 'Saldo' in df_anterior.columns else 0
    
        for _, aluno in alunos.iterrows():
            nome = aluno['Nome']
            turma = aluno['Turma']
            email = aluno['Email']
            contribuinte = aluno['Contribuinte']
            associado = aluno['Associado']
            
            if 'Contribuinte' in df_anterior.columns and 'Saldo' in df_anterior.columns:
                saldo_anterior = df_anterior.loc[df_anterior['Contribuinte'] == contribuinte, 'Saldo']
                saldo_anterior = saldo_anterior.values[0] if not saldo_anterior.empty else 0
            else:
                saldo_anterior = 0
        
            nr_acolhimento = calcular_nr_dias_acolhimento(contribuinte, caf_acolhimento)
            nr_prolongamento = calcular_nr_dias_prolongamento(contribuinte, caf_prolongamento)
        
            preco_caf = calcular_preco_caf(contribuinte, mes, caf_acolhimento, caf_prolongamento, precos, associado)
            preco_danca = calcular_preco_danca(contribuinte, danca, precos, mes, associado)
            preco_lanche = calcular_preco_lanche(contribuinte, lanche, precos, mes, associado)
            preco_karate = calcular_preco_karate(contribuinte, karate, precos, mes, associado)
            # Novo cálculo do valor recebido numerário
            valor_recebido_num = obter_valor_recebido_numerario(contribuinte, recebimentos)
            
            valor_recebido = obter_valor_recebido_transf(contribuinte, recebimentos_transf)
            recibo = ''
            saldo_formula = f"=J{len(dados_saida) + 2} + K{len(dados_saida) + 2}   +  L{len(dados_saida) + 2} - (G{len(dados_saida) + 2} + H{len(dados_saida) + 2} + I{len(dados_saida) + 2} + J{len(dados_saida) + 2} + Q{len(dados_saida) + 2} )"
        
            dados_saida.append([nome, turma, associado, contribuinte, nr_acolhimento, nr_prolongamento, preco_caf, preco_danca, preco_lanche, preco_karate, valor_recebido_num, valor_recebido, saldo_anterior, saldo_formula, recibo, email])

        df_saida = pd.DataFrame(dados_saida, columns=[
            'Nome', 'Turma', 'Associado', 'Contribuinte', 'Nr Acolhimento', 'Nr Prolongamento', 'Preco CAF', 'Preco Danca', 'Preco Lanche', 'Preço Karate', 'Valor Recebido Num', 'Valor Recebido Transf', 'Saldo Anterior', 'Saldo', 'Recibo', 'Email', 'Notas', 'ADICIONAIS'
        ])
    
        caminho_relatorio = os.path.join(mes, f'relatorioMensal_{mes}.xlsx')
    
        # Exportar para Excel e aplicar formatações
        with pd.ExcelWriter(caminho_relatorio, engine='openpyxl') as writer:
            df_saida.to_excel(writer, index=False, sheet_name='relatorioMensal')
            
            workbook = writer.book
            worksheet = writer.sheets['relatorioMensal']
    
            # Definir larguras das colunas
            column_widths = {
                'A':  42.0,  # 1 Nome
                'B':  12.3,  # 2 Turma
                'C':  15.0,  # 3 Associado
                'D':  12.3,  # 4 Contribuinte
                'E':  12.3,  # 5 Nr Acolhimento
                'F':  12.3,  # 6 Nr Prolongamento
                'G':  12.3,  # 7 Preco CAF
                'H':  12.3,  # 8 Preco Danca
                'I':  12.3,  # 9 Preco Lanche
                'J':  12.3,  # 10 Preco Karate
                'K':  12.3,  # 11 Valor Recebido
                'L':  12.3,  # 12 Valor Recebido
                'M':  12.3,  # 13 Saldo Anterior
                'N':  12.3,  # 14 Saldo
                'O':  12.3,  # 15 Recibo
                'P':  12.3,  # 16 Email
                'Q':  12.3   # 17 Adicionais
            }
    
            # Aplicar larguras de coluna
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
    
            # Ocultar colunas específicas
            # cols_to_hide = ['D', 'E', 'F', 'N']  # Colunas que você deseja ocultar
            # for col in cols_to_hide:
            #     worksheet.column_dimensions[col].hidden = True
            
            # Definir altura das linhas
            worksheet.row_dimensions[1].height = 46.5  # Largura da linha de cabeçalho
            for row in range(2, len(dados_saida) + 2):
                worksheet.row_dimensions[row].height = 15  # Largura das outras linhas
    
            # Formatação de "wrap text" na primeira linha (cabeçalho)
            for cell in worksheet[1]:  # Linha do cabeçalho
                cell.alignment = cell.alignment.copy(wrap_text=True)
        
            # Formatação de moeda para as colunas especificadas
            colunas_moeda = [7, 8, 9, 10, 11, 12, 13, 14]  # Índices das colunas para PrecoCAF, PrecoDanca, etc.
        
            for col in colunas_moeda:
                for row in range(2, len(dados_saida) + 2):
                    cell = worksheet.cell(row=row, column=col)
                    cell.number_format = '€ #,##0.00'  # Formato de moeda em euros
        
            # Formatação condicional para a coluna "Saldo"
            for row in range(2, len(dados_saida) + 2):
                saldo_cell = worksheet.cell(row=row, column=14)  # Coluna "Saldo"
                saldo_cell.font = Font(color="0000FF")  # Padrão azul
                
                saldoAnterior_cell = worksheet.cell(row=row, column=13)  # Coluna "SaldoAnterior"
                saldoAnterior_cell.font = Font(color="0000FF")  # Padrão azul
            
                # Adicionar formatação condicional para a coluna "Saldo"
                worksheet.conditional_formatting.add(
                    f'N{row}',  # Coluna "Saldo" (M)
                    CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=Font(color='FF0000'))  # Vermelho se menor que zero
                )
            
                # Adicionar formatação condicional para a coluna "Saldo Anterior"
                worksheet.conditional_formatting.add(
                    f'M{row}',  # Coluna "Saldo Anterior" (L)
                    CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=Font(color='FF0000'))  # Vermelho se menor que zero
                )
    
        # Backup do relatório
        if os.path.exists(caminho_relatorio):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            caminho_backup = os.path.join(mes, f'relatorioMensal_backup_{mes}_{timestamp}.xlsx')
            shutil.copy(caminho_relatorio, caminho_backup)
    except Exception as e:
        # Captura qualquer erro e imprime para ajudar na depuração
        print(f"Erro ao gerar relatório para o mês de {mes}: {str(e)}")


def obter_mes_anterior(mes):
    meses = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
    try:
        mes_index = meses.index(mes.strip().lower())
        if mes_index == 0:
            return meses[-1]  # Retorna dezembro se for janeiro
        return meses[mes_index - 1]
    except ValueError:
        return None



# Função principal
def main():
    meses = [
        "janeiro", "fevereiro", "março", "abril",
        "maio", "junho", "julho", "agosto",
        "setembro", "outubro", "novembro", "dezembro"
    ]
    
    print("##### ASSOCIAÇÃO DE PAIS ESCOLA DA FEIRA NOVA #####")
    print("Gerador de Relatório Mensal")
    print("---------------------")
    print("Escolha um mês para gerar o Relatório Mensal:")

    for i, mes in enumerate(meses, start=1):
        print(f"{i}. {mes.capitalize()}")

    print("0. Sair")

    while True:
        try:
            # Receber input e verificar se é um número inteiro válido
            escolha = input("Digite o número do mês (0 para sair): ").strip()
            print(f"Escolha digitada: {escolha}")  # Debug: Imprime o valor digitado

            # Verifique se o valor inserido é um número e converta
            if escolha.isdigit():
                escolha = int(escolha)
                
                if escolha == 0:
                    print("Obrigado e Bom Trabalho...")
                    break  # Sai do loop e encerra a execução do programa
                    
                elif 1 <= escolha <= 12:
                    mes_selecionado = meses[escolha - 1]  # Obtém o mês selecionado com base no número
                    print(f"Gerando relatório para o mês de {mes_selecionado.capitalize()}...")
                    gerar_relatorioMensal(mes_selecionado)  # Chama a função para gerar o relatório
                    print(f"Relatório Mensal gerado para o mês de {mes_selecionado.capitalize()} com sucesso!")
                    
                else:
                    print("Opção inválida. Tente novamente.")  # Opção fora do intervalo de 1 a 12
                
            else:
                print("Por favor, insira um número válido.")  # Captura erro se não for um número
                
        except ValueError as e:
            print("Erro no processamento do número. Tente novamente.")  # Captura erro de tipo de entrada


# Chamada da função principal
if __name__ == "__main__":
    main()
